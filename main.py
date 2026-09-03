"""
Despacho de Canales — Colbeef
Backend FastAPI con conexión directa a PostgreSQL (solo lectura)
v1.0 — Medias canales: Media Canal 1 (sufijo -1001) y Media Canal 2 (sufijo -1002)
"""
from fastapi import FastAPI, HTTPException, Header, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import os
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime
from io import BytesIO
from threading import Lock
from typing import Optional, List
from urllib.parse import quote

from dotenv import load_dotenv
import psycopg2
import psycopg2.extras
from psycopg2.pool import ThreadedConnectionPool
import apps_script_local
import usability

load_dotenv()

app = FastAPI(title="Despacho de Canales Colbeef", version="1.0")

# ─── IDs tipo_parte_producto para canales (verificados en BD sirt) ──
ID_MC1 = 4   # "Media Canal 1"
ID_MC2 = 5   # "Media Canal 2 Cola"
IDS_CANAL = (ID_MC1, ID_MC2)

# Turno según weekday de la fecha (Python: lunes=0 ... domingo=6)
TURNO_POR_WEEKDAY = {0: "LxM", 1: "MxM", 2: "MxJ", 3: "JxV", 4: "VxS", 5: "SxD", 6: "DxL"}


def turno_de_fecha(fecha_str: Optional[str] = None) -> str:
    d = date.fromisoformat(fecha_str) if fecha_str else date.today()
    return TURNO_POR_WEEKDAY[d.weekday()]


def resolver_turno(fecha_str: Optional[str], turno: Optional[str]) -> Optional[str]:
    """Si turno viene vacío o None, usa el de la fecha. 'Todos' = sin filtro."""
    if turno is None:
        return turno_de_fecha(fecha_str)
    t = str(turno).strip()
    if t.lower() == "todos":
        return None
    if t == "":
        return turno_de_fecha(fecha_str)
    return t


TURNOS_CODIGO = ("DxL", "LxM", "MxM", "MxJ", "JxV", "VxS", "SxD")

# Destinos/puestos de stock interno (no despacho a plaza), igual idea que Gestor Vísceras
PUESTOS_EXCLUIDOS = {
    "01305", "03105", "05200", "12157", "379P",
    "CAVA AJR", "CAVA FORTUNATO", "CAVA MIREYA", "CAVA.", "CAVA",
    "CCARNES CAVA", "OLIMPICA", "RH32", "DRA CAVA", "CAVA WO",
    "CAVAYERSON", "CAVA JUDITH", "CAVA CV", "CAVA EMERGENCIA",
}

SQL_JOINS_LOGISTICA = """
        LEFT JOIN trazabilidad_proceso.producto_empresa pe
            ON pe.id_producto::text = pp.id_producto::text AND pe.activo = true
        LEFT JOIN organizaciones.empresa e3
            ON e3.id = pe.id_empresa
        LEFT JOIN trazabilidad_proceso.parte_producto_empresa ppe
            ON ppe.id_producto::text = pp.id_producto::text AND ppe.id_parte_producto = pp.id
        LEFT JOIN trazabilidad_proceso.parte_producto_empresa_local ppel
            ON ppel.id_parte_producto_empresa = ppe.id
        LEFT JOIN organizaciones.sucursal s
            ON s.id = ppel.id_local
        LEFT JOIN trazabilidad_proceso.destino de
            ON de.id = s.id_destino
"""


def patron_turno(turno: Optional[str]) -> Optional[str]:
    """Patrón ILIKE si el texto trae el código (ej. /JxV/). None = sin filtro."""
    return f"%{turno}%" if turno else None


def formatear_codigo_sucursal(codigo) -> str:
    """09404 → 9404 (como Gestor Vísceras)."""
    raw = str(codigo or "").strip()
    if not raw:
        return ""
    if raw.isdigit():
        try:
            return str(int(raw))
        except ValueError:
            return raw
    return raw


def parse_puesto_operacion(puesto_full: str) -> dict:
    """
    Descompone ruta SIRT:
    09404/Floridablanca/PLAZA DE MERCADO... /JxV/
    → puesto 9404, zona Floridablanca, dirección, turno JxV
    """
    raw = str(puesto_full or "").strip()
    parts = [p.strip() for p in raw.split("/") if p.strip()]
    sin_turno = [p for p in parts if p not in TURNOS_CODIGO]
    codigo = formatear_codigo_sucursal(sin_turno[0] if sin_turno else "")
    zona = sin_turno[1].strip() if len(sin_turno) > 1 else ""
    direccion = sin_turno[2] if len(sin_turno) > 2 else ""
    turno = next((p for p in parts if p in TURNOS_CODIGO), "")
    etiqueta = f"{codigo} · {zona}" if codigo and zona else (codigo or zona or raw[:96])
    zona_key = zona.upper()
    if codigo and zona_key:
        clave = f"{str(codigo).upper()}|{zona_key}"
    elif codigo:
        clave = str(codigo).upper()
    else:
        clave = " ".join(raw.split()).upper()
    return {
        "codigo": codigo,
        "zona": zona,
        "direccion": direccion,
        "turno": turno,
        "etiqueta": etiqueta,
        "ruta": " / ".join(sin_turno),
        "ruta_completa": raw,
        "clave": clave,
    }


def construir_ruta(puesto: str, zona: str, direccion: str = "", turno: str = "") -> str:
    bits = [str(puesto or "").strip(), str(zona or "").strip(), str(direccion or "").strip()]
    bits = [b for b in bits if b]
    base = "/".join(bits)
    if turno:
        base = f"{base}/{turno}" if base else str(turno)
    return f"{base}/" if base else ""


def _ruta_cruda_desde_campos(con_destino: str, observaciones: str) -> str:
    """Prioriza con_destino u observaciones si ya traen la ruta con / (estilo Vísceras)."""
    for cand in (con_destino, observaciones):
        t = str(cand or "").strip()
        if not t or t.upper() in ("S", "N", "SI", "NO"):
            continue
        if "/" in t:
            return t
    return ""


def resolver_logistica_pieza(row: dict, turno_calendario: Optional[str] = None) -> dict:
    """Arma puesto/zona/turno como el Gestor de Vísceras."""
    con_dest = str(row.get("destino") or row.get("con_destino") or "").strip()
    obs = str(row.get("observaciones") or "").strip()
    suc = str(row.get("sucursal_origen") or row.get("sucursal") or "").strip()
    zona_db = str(row.get("destino_real") or row.get("zona") or "").strip()
    dir_db = str(row.get("direccion_entrega") or row.get("direccion") or "").strip()

    ruta_cruda = _ruta_cruda_desde_campos(con_dest, obs)
    if ruta_cruda:
        po = parse_puesto_operacion(ruta_cruda)
        puesto = po["codigo"] or formatear_codigo_sucursal(suc)
        zona = po["zona"] or zona_db
        direccion = po["direccion"] or dir_db
        turno = po["turno"] or (turno_calendario or "")
    else:
        puesto = formatear_codigo_sucursal(suc) or suc
        zona = zona_db
        direccion = dir_db
        turno = turno_calendario or ""

    ruta = construir_ruta(puesto, zona, direccion, turno)
    po2 = parse_puesto_operacion(ruta)
    return {
        "puesto": puesto,
        "zona": zona,
        "direccion": direccion,
        "turno_ruta": turno,
        "ruta": ruta,
        "etiqueta": po2["etiqueta"] or (f"{puesto} · {zona}" if puesto or zona else "Sin ruta"),
        "clave": po2["clave"] or "SIN RUTA",
    }


def es_destino_despacho(log: dict) -> bool:
    """Excluye stock de cava / puestos internos."""
    zona = str(log.get("zona") or "").strip().upper()
    puesto = str(log.get("puesto") or "").strip().upper()
    if not zona and not puesto:
        return False
    if zona in ("CAVA", "PLANTA"):
        return False
    if puesto.startswith("CAVA") or puesto in {p.upper() for p in PUESTOS_EXCLUIDOS}:
        return False
    return True


def pasa_filtro_turno(log: dict, turno: Optional[str]) -> bool:
    if not turno:
        return True
    return str(log.get("turno_ruta") or "").upper() == str(turno).upper()


def enriquecer_logistica(rows: List[dict], fecha_filtro: str, turno: Optional[str], solo_despacho: bool = True) -> List[dict]:
    """Aplica parseo de ruta + filtro de turno (calendario si la ruta no trae código)."""
    cal = turno or turno_de_fecha(fecha_filtro)
    out = []
    for r in rows:
        log = resolver_logistica_pieza(r, cal)
        if solo_despacho and not es_destino_despacho(log):
            continue
        if not pasa_filtro_turno(log, turno):
            continue
        r.update(log)
        # Destino visible = zona (Floridablanca), no la bandera "S"
        r["destino"] = log["zona"] or log["ruta"] or r.get("destino")
        out.append(r)
    return out


def sql_filtro_turno(col: str = "pp.con_destino") -> str:
    """Filtro SQL legacy (ILIKE). Preferir enriquecer_logistica en despacho/planilla."""
    return f"(%s::text IS NULL OR {col} ILIKE %s::text)"

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

DB_CONFIG = {
    "host":            os.getenv("POSTGRES_HOST", "10.64.1.47"),
    "port":            int(os.getenv("POSTGRES_PORT", "5432")),
    "dbname":          os.getenv("POSTGRES_DB", "sirt"),
    "user":            os.getenv("POSTGRES_USER", "acceso"),
    "password":        os.getenv("POSTGRES_PASSWORD", ""),
    "connect_timeout": int(os.getenv("POSTGRES_CONNECT_TIMEOUT", "5")),
    "options": f"-c statement_timeout={os.getenv('POSTGRES_STATEMENT_TIMEOUT_MS', '20000')}",
}

_POOL = None
_POOL_LOCK = Lock()
_CACHE = {}
_CACHE_LOCK = Lock()
CACHE_TTL_SEG = 12


def get_pool():
    global _POOL
    with _POOL_LOCK:
        if _POOL is None:
            _POOL = ThreadedConnectionPool(2, 10, **DB_CONFIG)
        return _POOL


def get_conn():
    try:
        return get_pool().getconn()
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"No se pudo conectar a la BD: {str(e)}")


def put_conn(conn):
    try:
        get_pool().putconn(conn)
    except Exception:
        try:
            conn.close()
        except Exception:
            pass


def query(sql: str, params=None):
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params or ())
            return [dict(r) for r in cur.fetchall()]
    finally:
        put_conn(conn)


def cache_get(key):
    with _CACHE_LOCK:
        hit = _CACHE.get(key)
        if hit and (time.time() - hit[0]) < CACHE_TTL_SEG:
            return hit[1]
    return None


def cache_set(key, value):
    with _CACHE_LOCK:
        _CACHE[key] = (time.time(), value)


def safe_query(sql: str, params=None, label: str = "consulta"):
    try:
        return query(sql, params)
    except Exception as e:
        print(f"[WARN] {label} falló: {e}")
        return []


def safe_query_many(tasks):
    if not tasks:
        return {}
    max_workers = min(len(tasks), 6)
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(safe_query, sql, params, label): key
            for key, sql, params, label in tasks
        }
        return {key: future.result() for future, key in futures.items()}


def serializable(rows):
    result = []
    for row in rows:
        clean = {}
        for k, v in row.items():
            if isinstance(v, (date, datetime)):
                clean[k] = v.isoformat()
            else:
                clean[k] = v
        result.append(clean)
    return result


def tipo_canal_label(id_tipo: int) -> str:
    if id_tipo == ID_MC1:
        return "Media Canal 1"
    if id_tipo == ID_MC2:
        return "Media Canal 2"
    return str(id_tipo)


def codigo_completo_canal(codigo: str, id_tipo: int) -> str:
    """
    Devuelve el código completo con sufijo numérico:
    Media Canal 1 → ...-1001
    Media Canal 2 → ...-1002
    Si la BD ya trae el sufijo, se respeta.
    """
    codigo = (codigo or "").strip()
    if not codigo:
        return ""
    partes = codigo.split("-")
    # Ya trae sufijo 1001/1002
    if len(partes) >= 3 and partes[-1] in ("1001", "1002", "001", "002"):
        return codigo
    sufijo_num = "1001" if id_tipo == ID_MC1 else ("1002" if id_tipo == ID_MC2 else "")
    if not sufijo_num:
        return codigo
    return f"{codigo}-{sufijo_num}"


def enriquecer_codigo(row: dict) -> dict:
    """Código completo tipo 2608-09418-1001 (nunca letras MC1/MC2)."""
    id_tipo = row.get("id_tipo", 0)
    codigo = row.get("codigo") or ""
    completo = codigo_completo_canal(codigo, id_tipo)
    row["sufijo"] = tipo_canal_label(id_tipo)
    row["codigo_completo"] = completo
    row["codigo_sufijo"] = completo
    row["codigo"] = completo  # el frontend muestra siempre el completo
    return row


def _nombres_opl_conocidos():
    cfg = apps_script_local.getOplConfig()
    names = {apps_script_local._as_str(o).upper() for o in (cfg.get("opls") or [])}
    names.add(apps_script_local.OPL_DEFAULT.upper())
    return names


def resolver_opl_de_propietario(propietario: str) -> str:
    prop = (propietario or "").strip()
    if not prop:
        return apps_script_local.OPL_DEFAULT
    conocidos = _nombres_opl_conocidos()
    prop_up = prop.upper()
    if prop_up in conocidos:
        return prop
    mapa = apps_script_local._opl_map(apps_script_local._load_state())
    return apps_script_local._resolver_opl(prop, mapa)


def consultar_canales_planilla(fecha_filtro: str, turno: Optional[str]):
    """
    Medias canales en cava con ruta logística (puesto/zona/turno),
    igual criterio que Gestor Vísceras: 09404/Floridablanca/.../JxV/
    """
    sql = f"""
        SELECT
            pp.id_producto                          AS codigo,
            tpp.id                                  AS id_tipo,
            COALESCE(NULLIF(TRIM(e3.nombre), ''), 'Sin propietario') AS propietario,
            c.nombre                                AS cava,
            r.nombre                                AS riel,
            pp.con_destino                          AS con_destino,
            pp.observaciones                        AS observaciones,
            s.nombre                                AS sucursal_origen,
            s.direccion                             AS direccion_entrega,
            de.nombre                               AS destino_real
        FROM trazabilidad_proceso.parte_producto pp
        JOIN trazabilidad_proceso.tipo_parte_producto tpp ON tpp.id = pp.id_tipo_parte_producto
        JOIN trazabilidad_proceso.parte_producto_cava_riel ppcr ON ppcr.id_parte_producto = pp.id
        LEFT JOIN trazabilidad_proceso.cava c ON c.id = ppcr.id_cava
        LEFT JOIN trazabilidad_proceso.riel r ON r.id = ppcr.id_riel
        {SQL_JOINS_LOGISTICA}
        WHERE pp.id_tipo_parte_producto IN %s
          AND ppcr.fecha_salida IS NULL
          AND ppcr.fecha_ingreso < (%s::date + INTERVAL '1 day')
          AND pp.con_destino IS NOT NULL
        ORDER BY e3.nombre NULLS LAST, c.orden NULLS LAST, r.nombre, pp.id_producto
    """
    rows = safe_query(sql, (IDS_CANAL, fecha_filtro), "planilla_puntos")
    data = enriquecer_logistica(serializable(rows), fecha_filtro, turno, solo_despacho=True)
    for r in data:
        enriquecer_codigo(r)
        r["opl"] = resolver_opl_de_propietario(r.get("propietario") or "")
        r["destino"] = r.get("zona") or ""
    data.sort(key=lambda x: (
        str(x.get("zona") or "").upper(),
        str(x.get("puesto") or ""),
        str(x.get("propietario") or ""),
        str(x.get("codigo") or ""),
    ))
    return data


def resumen_planilla_puntos(items: List[dict]):
    por_opl = {}
    for r in items:
        opl = r.get("opl") or apps_script_local.OPL_DEFAULT
        bucket = por_opl.setdefault(opl, {"opl": opl, "mc1": 0, "mc2": 0, "total_medias": 0})
        if r.get("id_tipo") == ID_MC1:
            bucket["mc1"] += 1
        elif r.get("id_tipo") == ID_MC2:
            bucket["mc2"] += 1
        bucket["total_medias"] += 1
    lista = []
    for bucket in por_opl.values():
        bucket["total_partes"] = bucket["total_medias"] * 0.5
        lista.append(bucket)
    lista.sort(key=lambda x: x["opl"])
    return lista


def construir_excel_opl(opl: str, fecha: str, turno, filas: List[dict]) -> BytesIO:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Falta openpyxl. En el servidor ejecuta: pip install openpyxl",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = (opl or "OPL")[:31]

    verde = PatternFill("solid", fgColor="259C39")
    verde_claro = PatternFill("solid", fgColor="E8F5E9")
    blanco = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    titulo = Font(color="FFFFFF", bold=True, name="Calibri", size=16)
    normal = Font(name="Calibri", size=11)
    thin = Border(
        left=Side(style="thin", color="C8E6C9"),
        right=Side(style="thin", color="C8E6C9"),
        top=Side(style="thin", color="C8E6C9"),
        bottom=Side(style="thin", color="C8E6C9"),
    )

    ws.merge_cells("A1:F1")
    ws["A1"] = f"OPL {opl}"
    ws["A1"].font = titulo
    ws["A1"].fill = verde
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for col in range(2, 7):
        ws.cell(1, col).fill = verde
    ws.row_dimensions[1].height = 28

    turno_txt = turno or "Todos"
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Medias canales pendientes · {fecha} · turno {turno_txt} · {len(filas)} registros"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="374151")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    headers = ["Código", "Propietario/Cliente", "Zona / Destino", "Puesto", "Cava", "Riel"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(3, i, h)
        cell.font = blanco
        cell.fill = verde
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin

    for idx, r in enumerate(filas, 4):
        vals = [
            r.get("codigo") or "",
            r.get("propietario") or "",
            r.get("zona") or r.get("destino") or "",
            r.get("puesto") or "",
            r.get("cava") or "",
            r.get("riel") or "",
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(idx, col, val)
            cell.font = normal
            cell.border = thin
            if idx % 2 == 0:
                cell.fill = verde_claro

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:F{max(3, 3 + len(filas))}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:3"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════
# PING
# ═══════════════════════════════════════════════════════
@app.get("/api/ping")
def ping():
    try:
        rows = query("SELECT current_database() AS db, now() AS ts")
        return {"ok": True, "db": rows[0]["db"], "ts": str(rows[0]["ts"])}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ═══════════════════════════════════════════════════════
# TIPOS DE CANAL — detección automática de IDs reales
# ═══════════════════════════════════════════════════════
@app.get("/api/diagnostico")
def get_diagnostico():
    """Diagnóstico ultraligero: muestra los últimos 5 registros de medias canales."""
    sql_sample = """
        SELECT id, id_producto, id_tipo_parte_producto, con_destino
        FROM trazabilidad_proceso.parte_producto
        WHERE id_tipo_parte_producto IN (4,5)
        LIMIT 5
    """
    return {
        "muestra_registros": serializable(safe_query(sql_sample, label="diag.sample")),
        "ids_buscados": [4, 5],
        "nombres": ["Media Canal 1", "Media Canal 2 Cola"],
    }


@app.get("/api/tipos_canal")
def get_tipos_canal():
    """Devuelve TODOS los tipos de parte_producto para identificar cuáles son canales."""
    sql = """
        SELECT id, nombre, abreviatura
        FROM trazabilidad_proceso.tipo_parte_producto
        ORDER BY id
    """
    rows = safe_query(sql, label="tipos_canal")
    return {"tipos": serializable(rows)}


# ═══════════════════════════════════════════════════════
# CAVAS — canales en cava a una fecha dada
# ═══════════════════════════════════════════════════════
@app.get("/api/cavas")
def get_cavas(fecha: Optional[str] = None, turno: Optional[str] = None):
    fecha_filtro = fecha or date.today().isoformat()
    turno = resolver_turno(fecha_filtro, turno)
    turno_filtro = None  # turno se aplica en logística (despachos/planilla); cavas = stock en cava
    sql = """
        SELECT
            pp.id                               AS id_parte_producto,
            pp.id_producto                      AS codigo,
            tpp.id                              AS id_tipo,
            tpp.nombre                          AS descripcion,
            tpp.abreviatura                     AS abrev_tipo,
            pp.identificacion                   AS identificacion,
            pp.con_destino                      AS destino,
            pp.observaciones                    AS observaciones,
            pp.reetiquetado,
            c.nombre                            AS cava,
            c.orden                             AS cava_orden,
            r.nombre                            AS riel,
            ppcr.fecha_ingreso                  AS fecha_ingreso_cava,
            ppcr.fecha_salida                   AS fecha_salida_cava,
            ppcr.numero_informacion_ingreso     AS numero_ingreso,
            p.peso_animal_pie                   AS peso_pie_kg,
            e3.nombre                           AS propietario,
            s.nombre                            AS sucursal_origen,
            de.nombre                           AS destino_real
        FROM trazabilidad_proceso.parte_producto pp
        JOIN trazabilidad_proceso.tipo_parte_producto tpp
            ON tpp.id = pp.id_tipo_parte_producto
        JOIN trazabilidad_proceso.parte_producto_cava_riel ppcr
            ON ppcr.id_parte_producto = pp.id
        LEFT JOIN trazabilidad_proceso.cava c
            ON c.id = ppcr.id_cava
        LEFT JOIN trazabilidad_proceso.riel r
            ON r.id = ppcr.id_riel
        LEFT JOIN trazabilidad_proceso.producto p
            ON p.id::text = pp.id_producto::text
        LEFT JOIN trazabilidad_proceso.producto_empresa pe
            ON pe.id_producto::text = p.id::text AND pe.activo = true
        LEFT JOIN organizaciones.empresa e3
            ON e3.id = pe.id_empresa
        LEFT JOIN trazabilidad_proceso.parte_producto_empresa ppe
            ON ppe.id_producto::text = pp.id_producto::text AND ppe.id_parte_producto = pp.id
        LEFT JOIN trazabilidad_proceso.parte_producto_empresa_local ppel
            ON ppel.id_parte_producto_empresa = ppe.id
        LEFT JOIN organizaciones.sucursal s
            ON s.id = ppel.id_local
        LEFT JOIN trazabilidad_proceso.destino de
            ON de.id = s.id_destino
        WHERE
            pp.id_tipo_parte_producto IN %s
            AND ppcr.fecha_salida IS NULL
            AND ppcr.fecha_ingreso < (%s::date + INTERVAL '1 day')
            AND """ + sql_filtro_turno() + """
        ORDER BY
            c.orden NULLS LAST,
            r.nombre,
            pp.id_producto
    """
    rows = safe_query(sql, (IDS_CANAL, fecha_filtro, turno_filtro, turno_filtro), "cavas")
    data = serializable(rows)
    for r in data:
        enriquecer_codigo(r)
    return {"fecha": fecha_filtro, "turno": turno, "total": len(data), "data": data}


# ═══════════════════════════════════════════════════════
# DASHBOARD — resumen ejecutivo de canales en cava
# ═══════════════════════════════════════════════════════
@app.get("/api/dashboard")
def get_dashboard(fecha: Optional[str] = None, turno: Optional[str] = None):
    fecha_filtro = fecha or date.today().isoformat()
    turno = resolver_turno(fecha_filtro, turno)
    turno_filtro = None  # turno se aplica en logística (despachos/planilla); cavas = stock en cava
    ck = ("dashboard", fecha_filtro, turno)
    hit = cache_get(ck)
    if hit is not None:
        return hit
    sql_totales = """
        SELECT
            COUNT(pp.id) FILTER (WHERE pp.id_tipo_parte_producto = %s) AS mc1,
            COUNT(pp.id) FILTER (WHERE pp.id_tipo_parte_producto = %s) AS mc2,
            COUNT(pp.id) AS total_partes,
            COUNT(DISTINCT SPLIT_PART(pp.id_producto, '-', 1)||'-'||SPLIT_PART(pp.id_producto, '-', 2)) AS animales
        FROM trazabilidad_proceso.parte_producto pp
        JOIN trazabilidad_proceso.parte_producto_cava_riel ppcr ON ppcr.id_parte_producto = pp.id
        WHERE pp.id_tipo_parte_producto IN %s
          AND ppcr.fecha_salida IS NULL
          AND ppcr.fecha_ingreso < (%s::date + INTERVAL '1 day')
          AND pp.con_destino IS NOT NULL
          AND """ + sql_filtro_turno() + """
    """
    sql_cavas = """
        SELECT c.nombre AS cava,
               COUNT(pp.id) FILTER (WHERE pp.id_tipo_parte_producto = %s) AS mc1,
               COUNT(pp.id) FILTER (WHERE pp.id_tipo_parte_producto = %s) AS mc2,
               COUNT(pp.id) AS total
        FROM trazabilidad_proceso.parte_producto pp
        JOIN trazabilidad_proceso.parte_producto_cava_riel ppcr ON ppcr.id_parte_producto = pp.id
        JOIN trazabilidad_proceso.cava c ON c.id = ppcr.id_cava
        WHERE pp.id_tipo_parte_producto IN %s
          AND ppcr.fecha_salida IS NULL
          AND ppcr.fecha_ingreso < (%s::date + INTERVAL '1 day')
          AND pp.con_destino IS NOT NULL
          AND """ + sql_filtro_turno() + """
        GROUP BY c.id, c.nombre, c.orden ORDER BY c.orden NULLS LAST
    """
    sql_destinos = """
        SELECT pp.con_destino AS destino, COUNT(pp.id) AS total
        FROM trazabilidad_proceso.parte_producto pp
        JOIN trazabilidad_proceso.parte_producto_cava_riel ppcr ON ppcr.id_parte_producto = pp.id
        WHERE pp.id_tipo_parte_producto IN %s
          AND ppcr.fecha_salida IS NULL
          AND ppcr.fecha_ingreso < (%s::date + INTERVAL '1 day')
          AND pp.con_destino IS NOT NULL
          AND """ + sql_filtro_turno() + """
        GROUP BY pp.con_destino ORDER BY total DESC LIMIT 10
    """

    results = safe_query_many([
        ("totales", sql_totales, (ID_MC1, ID_MC2, IDS_CANAL, fecha_filtro, turno_filtro, turno_filtro), "dashboard.totales"),
        ("cavas",   sql_cavas,   (ID_MC1, ID_MC2, IDS_CANAL, fecha_filtro, turno_filtro, turno_filtro), "dashboard.cavas"),
        ("destinos",sql_destinos,(IDS_CANAL, fecha_filtro, turno_filtro, turno_filtro),                 "dashboard.destinos"),
    ])
    t  = results.get("totales", [{}])
    cv = results.get("cavas", [])
    ds = results.get("destinos", [])

    tot = t[0] if t else {}
    mc1 = int(tot.get("mc1") or 0)
    mc2 = int(tot.get("mc2") or 0)
    # Un animal completo = MC1 + MC2 → cada media vale 0.5 canales
    canales_completas = (mc1 + mc2) / 2

    payload = {
        "fecha":             fecha_filtro,
        "turno":             turno or "Todos",
        "mc1":               mc1,
        "mc2":               mc2,
        "total_partes":      int(tot.get("total_partes") or 0),
        "animales_distintos":int(tot.get("animales") or 0),
        "canales_completas": canales_completas,
        "cavas":             serializable(cv),
        "top_destinos":      serializable(ds),
    }
    cache_set(ck, payload)
    return payload


# ═══════════════════════════════════════════════════════
# DESPACHOS — agrupado por puesto/zona (como Gestor Vísceras)
# Ruta: 09404/Floridablanca/.../JxV/  → puesto 9404, zona Floridablanca
# ═══════════════════════════════════════════════════════
@app.get("/api/despachos")
def get_despachos(fecha: Optional[str] = None, turno: Optional[str] = None):
    fecha_filtro = fecha or date.today().isoformat()
    turno = resolver_turno(fecha_filtro, turno)
    ck = ("despachos", fecha_filtro, turno, "ruta_v2")
    hit = cache_get(ck)
    if hit is not None:
        return hit

    sql = f"""
        SELECT
            pp.id_producto AS codigo,
            pp.id_tipo_parte_producto AS id_tipo,
            COALESCE(NULLIF(TRIM(e3.nombre), ''), 'Sin propietario') AS propietario,
            pp.con_destino AS con_destino,
            pp.observaciones AS observaciones,
            s.nombre AS sucursal_origen,
            s.direccion AS direccion_entrega,
            de.nombre AS destino_real
        FROM trazabilidad_proceso.parte_producto pp
        JOIN trazabilidad_proceso.parte_producto_cava_riel ppcr
            ON ppcr.id_parte_producto = pp.id
        {SQL_JOINS_LOGISTICA}
        WHERE
            pp.id_tipo_parte_producto IN %s
            AND ppcr.fecha_salida IS NULL
            AND ppcr.fecha_ingreso < (%s::date + INTERVAL '1 day')
            AND pp.con_destino IS NOT NULL
    """
    rows = safe_query(sql, (IDS_CANAL, fecha_filtro), "despachos")
    piezas = enriquecer_logistica(serializable(rows), fecha_filtro, turno, solo_despacho=True)

    grupos = {}
    for r in piezas:
        clave = r.get("clave") or "SIN RUTA"
        g = grupos.get(clave)
        if not g:
            g = {
                "clave": clave,
                "puesto": r.get("puesto") or "",
                "zona": r.get("zona") or "",
                "direccion": r.get("direccion") or "",
                "ruta": r.get("ruta") or "",
                "etiqueta": r.get("etiqueta") or "",
                "destino": r.get("zona") or r.get("ruta") or "",
                "propietario": r.get("propietario") or "Sin propietario",
                "codigo": r.get("codigo"),
                "id_tipo": r.get("id_tipo"),
                "mc1": 0,
                "mc2": 0,
                "total_medias": 0,
                "props": {},
            }
            grupos[clave] = g
        if r.get("id_tipo") == ID_MC1:
            g["mc1"] += 1
        elif r.get("id_tipo") == ID_MC2:
            g["mc2"] += 1
        g["total_medias"] += 1
        prop = r.get("propietario") or "Sin propietario"
        g["props"][prop] = g["props"].get(prop, 0) + 1
        if not g.get("codigo"):
            g["codigo"] = r.get("codigo")
            g["id_tipo"] = r.get("id_tipo")

    data = []
    for g in grupos.values():
        # Propietario más frecuente en el puesto
        if g["props"]:
            g["propietario"] = max(g["props"].items(), key=lambda kv: kv[1])[0]
        g["total_partes"] = g["total_medias"] * 0.5
        g["total_canales"] = g["total_partes"]
        g["mas_de_una"] = g["total_medias"] > 1
        g["clientes"] = len(g["props"])
        enriquecer_codigo(g)
        del g["props"]
        data.append(g)

    data.sort(key=lambda x: (str(x.get("zona") or "").upper(), str(x.get("puesto") or "")))
    total_partes = sum(float(r.get("total_partes") or 0) for r in data)
    totales = {
        "mc1":           sum(r.get("mc1", 0) or 0 for r in data),
        "mc2":           sum(r.get("mc2", 0) or 0 for r in data),
        "total_partes":  total_partes,
        "total_canales": total_partes,
        "clientes":      len(data),
        "puestos":       len(data),
    }
    payload = {"fecha": fecha_filtro, "turno": turno, "totales": totales, "data": data}
    cache_set(ck, payload)
    return payload


# ═══════════════════════════════════════════════════════
# DETALLE DESPACHO — por puesto/zona (clave) o propietario
# ═══════════════════════════════════════════════════════
@app.get("/api/despachos/detalle")
def get_despacho_detalle(
    propietario: Optional[str] = None,
    destino: Optional[str] = None,
    puesto: Optional[str] = None,
    zona: Optional[str] = None,
    clave: Optional[str] = None,
    fecha: Optional[str] = None,
    turno: Optional[str] = None,
):
    fecha_filtro = fecha or date.today().isoformat()
    turno = resolver_turno(fecha_filtro, turno)
    sql = f"""
        SELECT
            pp.id_producto                          AS codigo,
            tpp.id                                  AS id_tipo,
            tpp.nombre                              AS descripcion,
            tpp.abreviatura                         AS abrev,
            pp.con_destino                          AS con_destino,
            pp.observaciones,
            COALESCE(NULLIF(TRIM(e3.nombre), ''), 'Sin propietario') AS propietario,
            c.nombre                                AS cava,
            r.nombre                                AS riel,
            ppcr.fecha_ingreso                      AS fecha_ingreso,
            ppcr.fecha_salida                       AS fecha_salida,
            s.nombre                                AS sucursal_origen,
            s.direccion                             AS direccion_entrega,
            de.nombre                               AS destino_real
        FROM trazabilidad_proceso.parte_producto pp
        JOIN trazabilidad_proceso.tipo_parte_producto tpp ON tpp.id = pp.id_tipo_parte_producto
        JOIN trazabilidad_proceso.parte_producto_cava_riel ppcr ON ppcr.id_parte_producto = pp.id
        LEFT JOIN trazabilidad_proceso.cava c ON c.id = ppcr.id_cava
        LEFT JOIN trazabilidad_proceso.riel r ON r.id = ppcr.id_riel
        {SQL_JOINS_LOGISTICA}
        WHERE pp.id_tipo_parte_producto IN %s
          AND ppcr.fecha_salida IS NULL
          AND ppcr.fecha_ingreso < (%s::date + INTERVAL '1 day')
          AND pp.con_destino IS NOT NULL
        ORDER BY c.orden NULLS LAST, r.nombre, pp.id_producto
    """
    rows = safe_query(sql, (IDS_CANAL, fecha_filtro), "despacho_detalle")
    data = enriquecer_logistica(serializable(rows), fecha_filtro, turno, solo_despacho=True)

    clave_q = (clave or "").strip()
    puesto_q = formatear_codigo_sucursal(puesto or "")
    zona_q = (zona or destino or "").strip().upper()
    prop_q = (propietario or "").strip()

    filtradas = []
    for r in data:
        if clave_q and r.get("clave") != clave_q:
            continue
        if puesto_q and formatear_codigo_sucursal(r.get("puesto")) != puesto_q:
            continue
        if zona_q and str(r.get("zona") or "").strip().upper() != zona_q:
            continue
        if prop_q and not clave_q and not puesto_q and str(r.get("propietario") or "").strip() != prop_q:
            continue
        enriquecer_codigo(r)
        r["destino"] = r.get("zona") or r.get("ruta") or ""
        r["total_partes"] = 0.5
        filtradas.append(r)

    etiqueta = filtradas[0]["etiqueta"] if filtradas else (clave_q or prop_q or zona_q or "—")
    return {
        "fecha": fecha_filtro,
        "turno": turno,
        "propietario": prop_q or (filtradas[0].get("propietario") if filtradas else ""),
        "puesto": puesto_q or (filtradas[0].get("puesto") if filtradas else ""),
        "zona": zona_q or (filtradas[0].get("zona") if filtradas else ""),
        "destino": etiqueta,
        "total": len(filtradas),
        "data": filtradas,
    }


# ═══════════════════════════════════════════════════════
# OPL — canales por operador logístico
# Cada canal vale 0.5 (MC1 o MC2 = 0.5; par completo = 1.0)
# ═══════════════════════════════════════════════════════
@app.get("/api/opl")
def get_opl(fecha: Optional[str] = None, turno: Optional[str] = None):
    fecha_filtro = fecha or date.today().isoformat()
    turno = resolver_turno(fecha_filtro, turno)
    turno_filtro = None  # turno se aplica en logística (despachos/planilla); cavas = stock en cava
    sql = """
        SELECT
            e3.nombre                           AS propietario,
            pp.con_destino                      AS destino,
            COUNT(pp.id) FILTER (WHERE pp.id_tipo_parte_producto = %s) AS mc1,
            COUNT(pp.id) FILTER (WHERE pp.id_tipo_parte_producto = %s) AS mc2,
            COUNT(pp.id) AS total_partes,
            COUNT(pp.id) * 0.5                  AS total_canales
        FROM trazabilidad_proceso.parte_producto pp
        JOIN trazabilidad_proceso.parte_producto_cava_riel ppcr ON ppcr.id_parte_producto = pp.id
        LEFT JOIN trazabilidad_proceso.producto_empresa pe ON pe.id_producto::text = pp.id_producto::text AND pe.activo = true
        LEFT JOIN organizaciones.empresa e3 ON e3.id = pe.id_empresa
        WHERE pp.id_tipo_parte_producto IN %s
          AND ppcr.fecha_salida IS NULL
          AND ppcr.fecha_ingreso < (%s::date + INTERVAL '1 day')
          AND """ + sql_filtro_turno() + """
        GROUP BY e3.nombre, pp.con_destino
        ORDER BY e3.nombre NULLS LAST, pp.con_destino
    """
    rows = safe_query(sql, (ID_MC1, ID_MC2, IDS_CANAL, fecha_filtro, turno_filtro, turno_filtro), "opl")
    data = serializable(rows)
    # Agrupar por propietario
    opls = {}
    for r in data:
        prop = r.get("propietario") or "SIN PROPIETARIO"
        if prop not in opls:
            opls[prop] = {"propietario": prop, "mc1": 0, "mc2": 0, "total_partes": 0, "total_canales": 0.0, "destinos": []}
        opls[prop]["mc1"]          += int(r.get("mc1") or 0)
        opls[prop]["mc2"]          += int(r.get("mc2") or 0)
        opls[prop]["total_partes"] += int(r.get("total_partes") or 0)
        opls[prop]["total_canales"] = opls[prop]["mc1"] * 0.5 + opls[prop]["mc2"] * 0.5
        if r.get("destino"):
            opls[prop]["destinos"].append(r["destino"])
    lista = sorted(opls.values(), key=lambda x: x["total_canales"], reverse=True)
    return {"fecha": fecha_filtro, "turno": turno, "total_opls": len(lista), "data": lista}


# ═══════════════════════════════════════════════════════
# OPL DETALLE — lista completa de canales para un propietario
# Columnas: código con sufijo, cliente/propietario, cava, riel, fecha salida
# ═══════════════════════════════════════════════════════
@app.get("/api/opl/detalle")
def get_opl_detalle(propietario: str, fecha: Optional[str] = None):
    fecha_filtro = fecha or date.today().isoformat()
    sql = """
        SELECT
            pp.id_producto                          AS codigo,
            tpp.id                                  AS id_tipo,
            tpp.nombre                              AS descripcion,
            pp.con_destino                          AS destino,
            pp.observaciones,
            e3.nombre                               AS propietario,
            c.nombre                                AS cava,
            r.nombre                                AS riel,
            ppcr.fecha_ingreso                      AS fecha_ingreso,
            ppcr.fecha_salida                       AS fecha_salida,
            EXTRACT(EPOCH FROM (NOW() - ppcr.fecha_ingreso))/3600 AS horas_en_cava
        FROM trazabilidad_proceso.parte_producto pp
        JOIN trazabilidad_proceso.tipo_parte_producto tpp ON tpp.id = pp.id_tipo_parte_producto
        JOIN trazabilidad_proceso.parte_producto_cava_riel ppcr ON ppcr.id_parte_producto = pp.id
        LEFT JOIN trazabilidad_proceso.cava c ON c.id = ppcr.id_cava
        LEFT JOIN trazabilidad_proceso.riel r ON r.id = ppcr.id_riel
        LEFT JOIN trazabilidad_proceso.producto p ON p.id::text = pp.id_producto::text
        LEFT JOIN trazabilidad_proceso.producto_empresa pe ON pe.id_producto::text = p.id::text AND pe.activo = true
        LEFT JOIN organizaciones.empresa e3 ON e3.id = pe.id_empresa
        LEFT JOIN trazabilidad_proceso.parte_producto_empresa ppe ON ppe.id_producto::text = pp.id_producto::text AND ppe.id_parte_producto = pp.id
        LEFT JOIN trazabilidad_proceso.parte_producto_empresa_local ppel ON ppel.id_parte_producto_empresa = ppe.id
        WHERE pp.id_tipo_parte_producto IN %s
          AND ppcr.fecha_salida IS NULL
          AND ppcr.fecha_ingreso < (%s::date + INTERVAL '1 day')
          AND e3.nombre = %s
        ORDER BY c.orden NULLS LAST, r.nombre, pp.id_producto
    """
    rows = safe_query(sql, (IDS_CANAL, fecha_filtro, propietario), "opl_detalle")
    data = serializable(rows)
    for r in data:
        enriquecer_codigo(r)
        r["horas_en_cava"] = round(float(r.get("horas_en_cava") or 0), 1)
    mc1 = sum(1 for r in data if r.get("id_tipo") == ID_MC1)
    mc2 = sum(1 for r in data if r.get("id_tipo") == ID_MC2)
    return {
        "fecha": fecha_filtro,
        "propietario": propietario,
        "mc1": mc1, "mc2": mc2,
        "total_partes": len(data),
        "total_canales": (mc1 + mc2) * 0.5,
        "data": data
    }


# ═══════════════════════════════════════════════════════
# SALIDAS — canales despachadas (pistoleadas) en el rango
# ═══════════════════════════════════════════════════════
@app.get("/api/salidas")
def get_salidas(fecha: Optional[str] = None, dias: int = 1, turno: Optional[str] = None):
    fecha_fin = fecha or date.today().isoformat()
    turno = resolver_turno(fecha_fin, turno)
    turno_filtro = None  # turno se aplica en logística (despachos/planilla); cavas = stock en cava
    sql = """
        SELECT
            pp.id_producto          AS codigo,
            tpp.id                  AS id_tipo,
            tpp.nombre              AS descripcion,
            pp.con_destino          AS destino,
            pp.observaciones,
            e3.nombre               AS propietario,
            c.nombre                AS cava,
            r.nombre                AS riel,
            ppcr.fecha_ingreso      AS ingreso_cava,
            ppcr.fecha_salida       AS fecha_salida,
            EXTRACT(EPOCH FROM (ppcr.fecha_salida - ppcr.fecha_ingreso))/3600 AS horas_en_cava
        FROM trazabilidad_proceso.parte_producto pp
        JOIN trazabilidad_proceso.tipo_parte_producto tpp ON tpp.id = pp.id_tipo_parte_producto
        JOIN trazabilidad_proceso.parte_producto_cava_riel ppcr ON ppcr.id_parte_producto = pp.id
        LEFT JOIN trazabilidad_proceso.cava c ON c.id = ppcr.id_cava
        LEFT JOIN trazabilidad_proceso.riel r ON r.id = ppcr.id_riel
        LEFT JOIN trazabilidad_proceso.producto p ON p.id::text = pp.id_producto::text
        LEFT JOIN trazabilidad_proceso.producto_empresa pe ON pe.id_producto::text = p.id::text AND pe.activo = true
        LEFT JOIN organizaciones.empresa e3 ON e3.id = pe.id_empresa
        LEFT JOIN trazabilidad_proceso.parte_producto_empresa ppe ON ppe.id_producto::text = pp.id_producto::text AND ppe.id_parte_producto = pp.id
        LEFT JOIN trazabilidad_proceso.parte_producto_empresa_local ppel ON ppel.id_parte_producto_empresa = ppe.id
        WHERE pp.id_tipo_parte_producto IN %s
          AND ppcr.fecha_salida IS NOT NULL
          AND ppcr.fecha_salida >= (%s::date - (%s * INTERVAL '1 day'))
          AND ppcr.fecha_salida < (%s::date + INTERVAL '1 day')
          AND """ + sql_filtro_turno() + """
        ORDER BY ppcr.fecha_salida DESC, pp.id_producto
    """
    rows = safe_query(sql, (IDS_CANAL, fecha_fin, dias, fecha_fin, turno_filtro, turno_filtro), "salidas")
    data = serializable(rows)
    for r in data:
        enriquecer_codigo(r)
        r["horas_en_cava"] = round(float(r.get("horas_en_cava") or 0), 1)
    mc1 = sum(1 for r in data if r.get("id_tipo") == ID_MC1)
    mc2 = sum(1 for r in data if r.get("id_tipo") == ID_MC2)
    return {
        "fecha": fecha_fin, "dias_rango": dias, "turno": turno,
        "mc1": mc1, "mc2": mc2,
        "total_partes": len(data),
        "total_canales": (mc1 + mc2) * 0.5,
        "data": data,
    }


# ═══════════════════════════════════════════════════════
# PLANILLA OPL — progreso de despacho por propietario
# ═══════════════════════════════════════════════════════
@app.get("/api/planilla_opl")
def get_planilla_opl(fecha: Optional[str] = None, turno: Optional[str] = None):
    """
    Progreso OPL del turno de la fecha:
    - Pendientes: canales aún en cava cuyo destino trae el turno (DxL, JxV, etc.)
    - Despachadas: canales con fecha_salida del día Y destino del mismo turno
    Así solo entra lo pistoleado de ese turno, no salidas de otros.
    """
    fecha_filtro = fecha or date.today().isoformat()
    turno = resolver_turno(fecha_filtro, turno)
    turno_filtro = None  # turno se aplica en logística (despachos/planilla); cavas = stock en cava
    ck = ("planilla_opl", fecha_filtro, turno)
    hit = cache_get(ck)
    if hit is not None:
        return hit

    sql_en_cava = """
        SELECT
            e3.nombre                           AS propietario,
            COUNT(pp.id) FILTER (WHERE pp.id_tipo_parte_producto = %s) AS mc1_pend,
            COUNT(pp.id) FILTER (WHERE pp.id_tipo_parte_producto = %s) AS mc2_pend,
            COUNT(pp.id)                        AS total_pend
        FROM trazabilidad_proceso.parte_producto pp
        JOIN trazabilidad_proceso.parte_producto_cava_riel ppcr ON ppcr.id_parte_producto = pp.id
        LEFT JOIN trazabilidad_proceso.producto_empresa pe ON pe.id_producto::text = pp.id_producto::text AND pe.activo = true
        LEFT JOIN organizaciones.empresa e3 ON e3.id = pe.id_empresa
        WHERE pp.id_tipo_parte_producto IN %s
          AND ppcr.fecha_salida IS NULL
          AND ppcr.fecha_ingreso < (%s::date + INTERVAL '1 day')
          AND """ + sql_filtro_turno() + """
        GROUP BY e3.nombre ORDER BY total_pend DESC
    """
    sql_salidas = """
        SELECT
            e3.nombre                           AS propietario,
            COUNT(pp.id) FILTER (WHERE pp.id_tipo_parte_producto = %s) AS mc1_sal,
            COUNT(pp.id) FILTER (WHERE pp.id_tipo_parte_producto = %s) AS mc2_sal,
            COUNT(pp.id)                        AS total_sal
        FROM trazabilidad_proceso.parte_producto pp
        JOIN trazabilidad_proceso.parte_producto_cava_riel ppcr ON ppcr.id_parte_producto = pp.id
        LEFT JOIN trazabilidad_proceso.producto_empresa pe ON pe.id_producto::text = pp.id_producto::text AND pe.activo = true
        LEFT JOIN organizaciones.empresa e3 ON e3.id = pe.id_empresa
        WHERE pp.id_tipo_parte_producto IN %s
          AND ppcr.fecha_salida IS NOT NULL
          AND DATE(ppcr.fecha_salida) = %s::date
          AND """ + sql_filtro_turno() + """
        GROUP BY e3.nombre
    """

    results = safe_query_many([
        ("en_cava", sql_en_cava, (ID_MC1, ID_MC2, IDS_CANAL, fecha_filtro, turno_filtro, turno_filtro), "planilla.cava"),
        ("salidas",  sql_salidas, (ID_MC1, ID_MC2, IDS_CANAL, fecha_filtro, turno_filtro, turno_filtro), "planilla.salidas"),
    ])

    idx_cava = {
        (r.get("propietario") or "SIN PROPIETARIO"): r
        for r in serializable(results.get("en_cava", []))
    }
    idx_salidas = {
        (r.get("propietario") or "SIN PROPIETARIO"): r
        for r in serializable(results.get("salidas", []))
    }

    lista = []
    for prop in sorted(set(idx_cava) | set(idx_salidas)):
        cava = idx_cava.get(prop, {})
        sal  = idx_salidas.get(prop, {})
        mc1_pend = int(cava.get("mc1_pend") or 0)
        mc2_pend = int(cava.get("mc2_pend") or 0)
        mc1_sal  = int(sal.get("mc1_sal") or 0)
        mc2_sal  = int(sal.get("mc2_sal") or 0)
        total_pend = mc1_pend + mc2_pend
        total_sal  = mc1_sal  + mc2_sal
        total_ini  = total_pend + total_sal
        pct = round((total_sal / total_ini) * 100) if total_ini else 0
        lista.append({
            "propietario": prop,
            "mc1_pendiente":  mc1_pend,
            "mc2_pendiente":  mc2_pend,
            "mc1_despachado": mc1_sal,
            "mc2_despachado": mc2_sal,
            "total_pendiente": total_pend,
            "total_despachado": total_sal,
            "canales_pendiente":  total_pend * 0.5,
            "canales_despachado": total_sal  * 0.5,
            "total_inicial":     total_ini,
            "progreso_pct":      pct,
        })

    lista.sort(key=lambda x: x["total_pendiente"], reverse=True)

    totales = {
        "mc1_pend":  sum(r["mc1_pendiente"]   for r in lista),
        "mc2_pend":  sum(r["mc2_pendiente"]   for r in lista),
        "mc1_sal":   sum(r["mc1_despachado"]  for r in lista),
        "mc2_sal":   sum(r["mc2_despachado"]  for r in lista),
        "pend_total":sum(r["total_pendiente"] for r in lista),
        "sal_total": sum(r["total_despachado"]for r in lista),
        "canales_pend": sum(r["canales_pendiente"]  for r in lista),
        "canales_sal":  sum(r["canales_despachado"] for r in lista),
    }
    t_ini = totales["pend_total"] + totales["sal_total"]
    totales["progreso_global"] = round((totales["sal_total"] / t_ini) * 100) if t_ini else 0

    payload = {
        "fecha": fecha_filtro,
        "turno": turno or turno_de_fecha(fecha_filtro),
        "totales": totales,
        "data": lista,
    }
    cache_set(ck, payload)
    return payload


# ═══════════════════════════════════════════════════════
# PLANILLA DE PUNTOS — lista por OPL para logística
# ═══════════════════════════════════════════════════════
def armar_planilla_estilo_visceras(items: List[dict], opl_sel: Optional[str], fecha: str, turno: Optional[str]) -> dict:
    """
    Misma estructura que Gestor Vísceras generarPlanillaPuntos:
    zonas[{nombre, total, puestos[{puesto, cantidad}]}] + lista plana puestos.
    Cantidad = medias × 0.5 (equivalente canal).
    """
    opl_sel = (opl_sel or "").strip()
    total_global = round(len(items) * 0.5, 2)
    zonas_map = {}
    puestos_flat = []
    total_opl = 0.0

    for r in items:
        opl_reg = r.get("opl") or apps_script_local.OPL_DEFAULT
        if opl_sel and opl_sel.upper() != "TODOS" and opl_reg != opl_sel:
            continue
        cantidad = 0.5
        total_opl += cantidad
        zona = (r.get("zona") or "SIN ZONA").strip() or "SIN ZONA"
        puesto = formatear_codigo_sucursal(r.get("puesto") or "") or "—"
        clave = r.get("clave") or f"{puesto}|{zona.upper()}"
        if zona not in zonas_map:
            zonas_map[zona] = {"total": 0.0, "puestos_map": {}}
        zonas_map[zona]["total"] += cantidad
        pm = zonas_map[zona]["puestos_map"]
        if clave not in pm:
            pm[clave] = {"puesto": puesto, "cantidad": 0.0}
        pm[clave]["cantidad"] += cantidad
        puestos_flat.append({
            "puesto": r.get("ruta") or construir_ruta(puesto, zona, r.get("direccion") or "", turno or ""),
            "etiqueta": r.get("etiqueta") or (f"{puesto} · {zona}" if puesto and zona else puesto or zona),
            "sucursal": puesto,
            "zona": zona,
            "cantidad": cantidad,
            "opl": opl_reg,
            "codigo": r.get("codigo"),
            "propietario": r.get("propietario"),
        })

    # Consolidar flat por puesto+zona
    flat_agg = {}
    for p in puestos_flat:
        k = f"{p['sucursal']}|{str(p['zona']).upper()}"
        if k not in flat_agg:
            flat_agg[k] = {
                "puesto": p["puesto"],
                "etiqueta": p["etiqueta"],
                "sucursal": p["sucursal"],
                "zona": p["zona"],
                "cantidad": 0.0,
                "opl": p["opl"],
            }
        flat_agg[k]["cantidad"] += p["cantidad"]
    puestos_lista = sorted(
        ({**v, "cantidad": round(v["cantidad"], 2)} for v in flat_agg.values()),
        key=lambda x: (str(x.get("zona") or ""), str(x.get("sucursal") or "")),
    )

    zonas_array = []
    for zona, bucket in zonas_map.items():
        puestos_arr = sorted(
            [
                {"puesto": v["puesto"], "cantidad": round(v["cantidad"], 2)}
                for v in bucket["puestos_map"].values()
            ],
            key=lambda x: str(x["puesto"]),
        )
        zonas_array.append({
            "nombre": zona,
            "total": round(bucket["total"], 2),
            "puestos": puestos_arr,
        })
    zonas_array.sort(key=lambda z: (-z["total"], z["nombre"]))

    pct = f"{(total_opl / total_global * 100):.1f}" if total_global > 0 else "0.0"
    return {
        "success": True,
        "opl": opl_sel or "TODOS",
        "zonas": zonas_array,
        "puestos": puestos_lista,
        "totalOPL": round(total_opl, 2),
        "totalGlobal": total_global,
        "porcentaje": pct,
        "turno": turno or "Todos",
        "fecha": fecha,
    }


@app.get("/api/planilla_puntos")
def get_planilla_puntos(
    fecha: Optional[str] = None,
    turno: Optional[str] = None,
    opl: Optional[str] = None,
):
    fecha_filtro = fecha or date.today().isoformat()
    turno = resolver_turno(fecha_filtro, turno)
    items = consultar_canales_planilla(fecha_filtro, turno)
    resumen = resumen_planilla_puntos(items)
    cfg = apps_script_local.getOplConfig()
    opls_cfg = cfg.get("opls") or []
    opls_data = [r["opl"] for r in resumen]
    opls = sorted(set(opls_cfg) | set(opls_data))
    pack = armar_planilla_estilo_visceras(items, opl, fecha_filtro, turno)
    pack["opls"] = opls
    pack["resumen"] = resumen
    # Detalle pieza a pieza (Excel / apoyo)
    opl_sel = (opl or "").strip()
    detalle = items if not opl_sel or opl_sel.upper() == "TODOS" else [
        r for r in items if (r.get("opl") or "") == opl_sel
    ]
    pack["data"] = detalle
    pack["total"] = len(detalle)
    pack["total_partes"] = round(len(detalle) * 0.5, 2)
    return pack


@app.get("/api/planilla_puntos/excel")
def excel_planilla_puntos(
    opl: str,
    fecha: Optional[str] = None,
    turno: Optional[str] = None,
):
    opl = (opl or "").strip()
    if not opl:
        raise HTTPException(status_code=400, detail="Indica el OPL")
    fecha_filtro = fecha or date.today().isoformat()
    turno = resolver_turno(fecha_filtro, turno)
    items = [
        r for r in consultar_canales_planilla(fecha_filtro, turno)
        if (r.get("opl") or "") == opl
    ]
    buf = construir_excel_opl(opl, fecha_filtro, turno, items)
    fname = f"OPL_{opl.replace(' ', '_')}_{fecha_filtro}.xlsx"
    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"
    }
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


# ═══════════════════════════════════════════════════════
# SERVIDOR
# ═══════════════════════════════════════════════════════
app.mount("/static", StaticFiles(directory="static"), name="static")


class AppsScriptRequest(BaseModel):
    args: list = []


class UsageEventIn(BaseModel):
    usuario: str = "anonimo"
    action: str = "event"
    module: str = ""
    detail: str = ""
    sessionId: str = ""
    page: str = ""
    meta: dict = {}


class UsageLoginIn(BaseModel):
    password: str = ""


@app.post("/api/usability/event")
def usability_event(payload: UsageEventIn, request: Request):
    ip = request.client.host if request.client else ""
    ua = request.headers.get("user-agent", "")
    return usability.record_event(payload.dict(), ip=ip, user_agent=ua)


@app.post("/api/usability/login")
def usability_login(payload: UsageLoginIn):
    token = usability.login_admin(payload.password)
    if not token:
        return {"success": False, "message": "Contraseña incorrecta"}
    return {"success": True, "token": token}


@app.get("/api/usability/stats")
def usability_stats(
    days: int = 30,
    x_usability_admin: str = Header(default="", alias="X-Usability-Admin"),
    authorization: str = Header(default=""),
):
    token = (x_usability_admin or "").strip()
    if not token and authorization.lower().startswith("bearer "):
        token = authorization[7:].strip()
    if not usability.verify_admin(token):
        raise HTTPException(status_code=401, detail="No autorizado")
    return usability.get_stats(days)


@app.get("/usabilidad.html")
def usabilidad_page():
    return FileResponse("static/usabilidad.html")


@app.post("/api/apps-script/{function_name}")
def run_apps_script_function(function_name: str, payload: AppsScriptRequest):
    try:
        return apps_script_local.dispatch(function_name, payload.args)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/favicon.ico")
def favicon():
    return FileResponse("static/favicon.svg", media_type="image/svg+xml")


@app.get("/")
def root():
    return FileResponse("static/index.html")


if __name__ == "__main__":
    import uvicorn
    host = os.getenv("APP_HOST", "0.0.0.0")
    port = int(os.getenv("APP_PORT", "8000"))
    uvicorn.run("main:app", host=host, port=port, reload=True)
